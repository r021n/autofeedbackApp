from django.shortcuts import render, redirect, get_object_or_404
from django.http import FileResponse, HttpResponse, HttpResponseServerError
from django.db import connection
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Topics, Questions, Answers1, Answers2, Answers3
from .forms import RegisterForm
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from collections import defaultdict
from django.core.paginator import Paginator

import re
import time
import google.generativeai as genai
import openpyxl

# =============================================================================================================================



# FUNCTION FOR SIMPLIFICATION
def getFeedback(answer):
    with open("appBelajar/api.txt") as my_file:
        API = my_file.read()
    genai.configure(api_key=API)

    # Create the model
    generation_config = {
    "temperature": 1,
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8192,
    "response_mime_type": "text/plain",
    }

    model = genai.GenerativeModel(
    model_name="tunedModels/trainingai3-desember-2024-3yfxz4du4z76",
    generation_config=generation_config,
    )

    chat_session = model.start_chat(
    history=[
    ]
    )

    response = chat_session.send_message(answer)

    print(response.text)

    return response.text
# =============================================================================================================================

# Create your views here.
def index(request):
    # print ("current position", os.getcwd())
    # print("File exists:", os.path.isfile('api.txt'))
    return render(request, 'appBelajar/index.html')


def about(request):
    return render(request, 'appBelajar/about.html')

def topics(request):
    q = request.GET.get('q', '')
    topics = Topics.objects.filter(Q(name__icontains=q)).order_by('-id')
    paginator = Paginator(topics, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'page_obj': page_obj}
    return render(request, 'appBelajar/topics.html', context)

def chance(request, pk):
    topic = Topics.objects.get(id=pk)
    totalQuestion = Questions.objects.filter(topic__id=pk)
    answered1 = Answers1.objects.filter(user=request.user, topic__id=pk)
    answered2 = Answers2.objects.filter(user=request.user, topic__id=pk)
    answered3 = Answers3.objects.filter(user=request.user, topic__id=pk)
    context = {'topic':topic, 'total':totalQuestion, 'answer1':answered1, 'answer2':answered2, 'answer3':answered3}
    return render(request, 'appBelajar/chance.html', context)

@user_passes_test(lambda u: u.is_superuser)
def topicList(request):
    q = request.GET.get('q', '')
    topics = Topics.objects.filter(Q(name__icontains=q)).order_by('-id')
    paginator = Paginator(topics, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {'page_obj': page_obj}
    return render(request, 'appBelajar/topicList.html', context)

@user_passes_test(lambda u: u.is_superuser)
def createTopic(request):
    if request.method == 'POST':
        topic = Topics.objects.create(name=request.POST['topic'])
        return redirect('questionList', topic.id)

    return render(request, 'appBelajar/createTopic.html')

@user_passes_test(lambda u: u.is_superuser)
def deleteTopic(request, pk):
    topic = get_object_or_404(Topics, id=pk)

    if request.method == 'POST':
        topic.delete()
        return redirect('topicList')

    return render(request, 'appBelajar/delete.html', {'obj': topic})

@user_passes_test(lambda u: u.is_superuser)
def questionList(request, pk):
    topic = get_object_or_404(Topics, id=pk)
    questions = Questions.objects.filter(topic=topic)

    if request.method == 'POST':
        question_body = request.POST['question']
        image_description = request.POST['imageDescription']
        image = request.FILES.get('image')

        question = Questions.objects.create(
            topic=topic,
            question=question_body,
            image=image,
            image_description=image_description
        )
        return redirect('questionList', pk=topic.pk)

    context = {'questions': questions, 'topic': topic}
    return render(request, 'appBelajar/questionList.html', context)

@user_passes_test(lambda u: u.is_superuser)
def updateQuestion(request, pk):
    question = get_object_or_404(Questions, id=pk)

    if request.method == 'POST':
        question_body = request.POST['question']
        image_description = request.POST['imageDescription']
        image = request.FILES.get('image')

        if image:
            question.image.delete()
            question.image = image

        question.question = question_body
        question.image_description = image_description
        question.save()
        return redirect('questionList', pk=question.topic.pk)

    return render(request, 'appBelajar/updateQuestion.html', {'question': question})

@user_passes_test(lambda u: u.is_superuser)
def deleteQuestion(request, pk):
    question = get_object_or_404(Questions, id=pk)

    if request.method == 'POST':
        topic_id = question.topic.pk
        question.delete()
        return redirect('questionList', pk=topic_id)

    return render(request, 'appBelajar/delete.html', {'obj': question})

@login_required(login_url='loginPage')
def exercise1(request, pk, number):
    topic = Topics.objects.get(id=pk)
    questions = Questions.objects.filter(topic=topic)
    question = questions[number - 1]

    def extract_score(text):
        patterns = [
            r"(?<=Skor:\s)(\d+)",
            r"(r'\d+(?!.*\d)')",
            r"(?i)skor\s*(\d+)"
        ]

        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1)

        return ""

    def extract_feedback(text):
        text = re.sub(r">", "", text)
        text = re.sub(r"Skor:\s*\d+", "", text)
        return text.strip()

    def removeStar(text):
        return re.sub(r"\*", "", text)

    if request.method == "POST":
        feedback = ""
        user_score = 0
        answer = request.POST.get('answer', "")

        if not answer.strip():
            feedback = "Kamu belum memasukkan jawaban"
        else:
            question_image = question.image_description
            prompt = f"gambar/grafik/tabel: {question_image}, soal: {question.question}, jawaban: {answer}, umpan balik dan skor antara 1 sampai 3 : (contoh output yang diharapkan = 'Jawbanmu ...(kurang tepat atau sudah benar, sesuaikan dengan konteks), (Coba perhatikan kembali...., seuaikan konteks). Skor: 3')"
            
            feedback_fix = removeStar(getFeedback(prompt))

            try:
                user_score = int(extract_score(feedback_fix))
            except ValueError:
                user_score = 0

            feedback = extract_feedback(feedback_fix)

        answer_obj, created = Answers1.objects.get_or_create(
            user=request.user,
            topic=topic,
            question=question,
            defaults={
                'answer': answer,
                'feedback': feedback,
                'score': user_score,
            }
        )

        if created:
            topic.participants.add(request.user)

    user_answer = Answers1.objects.filter(topic=topic, user=request.user, question=question).first()
    context = {
        'questions': questions,
        'questions_length': len(questions),
        'question': question,
        'number': number,
        'user_answer': user_answer
    }
    return render(request, 'appBelajar/exercise1.html', context)

@login_required(login_url='loginPage')
def exercise2(request, pk, number):
    topic = Topics.objects.get(id=pk)
    questions = Questions.objects.filter(topic=topic)
    question = questions[number - 1]

    def extract_score(text):
        patterns = [
            r"(?<=Skor:\s)(\d+)",
            r"(r'\d+(?!.*\d)')",
            r"(?i)skor\s*(\d+)"
        ]

        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1)

        return ""

    def extract_feedback(text):
        text = re.sub(r">", "", text)
        text = re.sub(r"Skor:\s*\d+", "", text)
        return text.strip()

    def removeStar(text):
        return re.sub(r"\*", "", text)

    if request.method == "POST":
        feedback = ""
        user_score = 0
        answer = request.POST.get('answer', "")

        if not answer.strip():
            feedback = "Kamu belum memasukkan jawaban"
        else:
            question_image = question.image_description
            prompt = f"gambar: {question_image}, soal: {question.question}, jawaban: {answer}, umpan balik dan skor antara 1 sampai 3 : (contoh output yang diharapkan = 'Jawbanmu ...(kurang tepat atau sudah benar, sesuaikan dengan konteks), (Coba perhatikan kembali...., seuaikan konteks). Skor: 3') "
            
            feedback_fix = removeStar(getFeedback(prompt))

            try:
                user_score = int(extract_score(feedback_fix))
            except ValueError:
                user_score = 0

            feedback = extract_feedback(feedback_fix)

        answer_obj, created = Answers2.objects.get_or_create(
            user=request.user,
            topic=topic,
            question=question,
            defaults={
                'answer': answer,
                'feedback': feedback,
                'score': user_score,
            }
        )

        if created:
            topic.participants.add(request.user)

    user_answer = Answers2.objects.filter(topic=topic, user=request.user, question=question).first()
    context = {
        'questions': questions,
        'questions_length': len(questions),
        'question': question,
        'number': number,
        'user_answer': user_answer
    }
    return render(request, 'appBelajar/exercise2.html', context)

@login_required(login_url='loginPage')
def exercise3(request, pk, number):
    topic = Topics.objects.get(id=pk)
    questions = Questions.objects.filter(topic=topic)
    question = questions[number - 1]

    def extract_score(text):
        patterns = [
            r"(?<=Skor:\s)(\d+)",
            r"(r'\d+(?!.*\d)')",
            r"(?i)skor\s*(\d+)"
        ]

        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1)

        return ""

    def extract_feedback(text):
        text = re.sub(r">", "", text)
        text = re.sub(r"Skor:\s*\d+", "", text)
        return text.strip()

    def removeStar(text):
        return re.sub(r"\*", "", text)

    if request.method == "POST":
        feedback = ""
        user_score = 0
        answer = request.POST.get('answer', "")

        if not answer.strip():
            feedback = "Kamu belum memasukkan jawaban"
        else:
            question_image = question.image_description
            prompt = f"gambar: {question_image}, soal: {question.question}, jawaban: {answer}, umpan balik dan skor antara 1 sampai 3 : (contoh output yang diharapkan = 'Jawbanmu ...(kurang tepat atau sudah benar, sesuaikan dengan konteks), (Coba perhatikan kembali...., seuaikan konteks). Skor: 3') "

            feedback_fix = removeStar(getFeedback(prompt))

            try:
                user_score = int(extract_score(feedback_fix))
            except ValueError:
                user_score = 0

            feedback = extract_feedback(feedback_fix)

        answer_obj, created = Answers3.objects.get_or_create(
            user=request.user,
            topic=topic,
            question=question,
            defaults={
                'answer': answer,
                'feedback': feedback,
                'score': user_score,
            }
        )

        if created:
            topic.participants.add(request.user)

    user_answer = Answers3.objects.filter(topic=topic, user=request.user, question=question).first()
    context = {
        'questions': questions,
        'questions_length': len(questions),
        'question': question,
        'number': number,
        'user_answer': user_answer
    }
    return render(request, 'appBelajar/exercise3.html', context)

@login_required(login_url='loginPage')
def exerciseOver1(request, pk):
    results = Answers1.objects.filter(user=request.user, topic__id=pk)

    score_sum = results.aggregate(Sum('score'))['score__sum'] or 0
    question_count = results.count()

    if question_count > 0:
        total_score = int((score_sum / (3 * question_count)) * 100)
    else:
        total_score = 0

    return render(request, 'appBelajar/over.html', {'score':total_score})

@login_required(login_url='loginPage')
def exerciseOver2(request, pk):
    results = Answers2.objects.filter(user=request.user, topic__id=pk)

    score_sum = results.aggregate(Sum('score'))['score__sum'] or 0
    question_count = results.count()

    if question_count > 0:
        total_score = int((score_sum / (3 * question_count)) * 100)
    else:
        total_score = 0

    return render(request, 'appBelajar/over.html', {'score':total_score})

@login_required(login_url='loginPage')
def exerciseOver3(request, pk):
    results = Answers3.objects.filter(user=request.user, topic__id=pk)

    score_sum = results.aggregate(Sum('score'))['score__sum'] or 0
    question_count = results.count()

    if question_count > 0:
        total_score = int((score_sum / (3 * question_count)) * 100)
    else:
        total_score = 0

    return render(request, 'appBelajar/over.html', {'score':total_score})

@login_required(login_url='loginPage')
def myResult1(request, pk):
    topic = Topics.objects.get(id=pk)
    results = Answers1.objects.filter(user=request.user, topic=topic)

    score_sum = results.aggregate(Sum('score'))['score__sum'] or 0
    question_count = Questions.objects.filter(topic__id=pk).count()

    if question_count > 0:
        total_score = int((score_sum / (3 * question_count)) * 100)
    else:
        total_score = 0

    context = {'results': results, 'score': total_score, 'topic': topic}
    return render(request, 'appBelajar/myResult1.html', context)

@login_required(login_url='loginPage')
def myResult2(request, pk):
    topic = Topics.objects.get(id=pk)
    results = Answers2.objects.filter(user=request.user, topic=topic)

    score_sum = results.aggregate(Sum('score'))['score__sum'] or 0
    question_count = Questions.objects.filter(topic__id=pk).count()

    if question_count > 0:
        total_score = int((score_sum / (3 * question_count)) * 100)
    else:
        total_score = 0

    context = {'results': results, 'score': total_score, 'topic': topic}
    return render(request, 'appBelajar/myResult2.html', context)

@login_required(login_url='loginPage')
def myResult3(request, pk):
    topic = Topics.objects.get(id=pk)
    results = Answers3.objects.filter(user=request.user, topic=topic)

    score_sum = results.aggregate(Sum('score'))['score__sum'] or 0
    question_count = Questions.objects.filter(topic__id=pk).count()

    if question_count > 0:
        total_score = int((score_sum / (3 * question_count)) * 100)
    else:
        total_score = 0

    context = {'results': results, 'score': total_score, 'topic': topic}
    return render(request, 'appBelajar/myResult3.html', context)

def registerPage(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('loginPage')
    else:
        form = RegisterForm()

    return render(request, 'appBelajar/register.html', {'form': form})

def loginPage(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Username atau password salah')
    else:
        try:
            User.objects.get(username=request.POST.get('username'))
        except User.DoesNotExist:
            messages.error(request, 'Akun tidak ditemukan')

    return render(request, 'appBelajar/loginForm.html')

@login_required(login_url='loginPage')
def logoutPage(request):
    logout(request)
    return redirect('home')

@user_passes_test(lambda u: u.is_superuser)
def studentsAnswer1(request, pk):
    topic = Topics.objects.get(id=pk)
    q = request.GET.get('q', '')
    filtered_answers = Answers1.objects.filter(Q(user__username__icontains=q), topic=topic).select_related('user', 'question')
    questions = list(Questions.objects.filter(topic=topic))

    def generate_header(questions):
        headers = ['nama siswa']
        for i, question in enumerate(questions, start=1):
            headers.append(f"soal nomer {i}: {question.question}")
            headers.append(f"Umpan Balik soal nomer {i}")
        headers.append('nilai siswa')
        return headers

    headers = generate_header(questions)

    user_answers = defaultdict(lambda: {'answers': {}, 'feedback': {}, 'score': {}})
    for answer in filtered_answers:
        user = answer.user.username
        question_text = answer.question.question
        user_answers[user]['answers'][question_text] = answer.answer
        user_answers[user]['feedback'][question_text] = answer.feedback
        user_answers[user]['score'][question_text] = answer.score or 0

    row_table = []
    for user, data in user_answers.items():
        row_data = [user]
        user_score = 0
        for question in questions:
            row_data.append(data['answers'].get(question.question, ''))
            row_data.append(data['feedback'].get(question.question, ''))
            user_score += data['score'].get(question.question, 0)
        final_score = int(user_score / (3 * len(questions)) * 100)
        row_data.append(final_score)
        row_table.append(row_data)

    context = {'headers': headers, 'row_data': row_table, 'topic': topic}
    return render(request, 'appBelajar/studentsAnswer1.html', context)

@user_passes_test(lambda u: u.is_superuser)
def studentsAnswer2(request, pk):
    topic = Topics.objects.get(id=pk)
    q = request.GET.get('q', '')
    filtered_answers = Answers2.objects.filter(Q(user__username__icontains=q), topic=topic).select_related('user', 'question')
    questions = list(Questions.objects.filter(topic=topic))

    def generate_header(questions):
        headers = ['nama siswa']
        for i, question in enumerate(questions, start=1):
            headers.append(f"soal nomer {i}: {question.question}")
            headers.append(f"Umpan Balik soal nomer {i}")
        headers.append('nilai siswa')
        return headers

    headers = generate_header(questions)

    user_answers = defaultdict(lambda: {'answers': {}, 'feedback': {}, 'score': {}})
    for answer in filtered_answers:
        user = answer.user.username
        question_text = answer.question.question
        user_answers[user]['answers'][question_text] = answer.answer
        user_answers[user]['feedback'][question_text] = answer.feedback
        user_answers[user]['score'][question_text] = answer.score or 0

    row_table = []
    for user, data in user_answers.items():
        row_data = [user]
        user_score = 0
        for question in questions:
            row_data.append(data['answers'].get(question.question, ''))
            row_data.append(data['feedback'].get(question.question, ''))
            user_score += data['score'].get(question.question, 0)
        final_score = int(user_score / (3 * len(questions)) * 100)
        row_data.append(final_score)
        row_table.append(row_data)

    context = {'headers': headers, 'row_data': row_table, 'topic': topic}
    return render(request, 'appBelajar/studentsAnswer2.html', context)

@user_passes_test(lambda u: u.is_superuser)
def studentsAnswer3(request, pk):
    topic = Topics.objects.get(id=pk)
    q = request.GET.get('q', '')
    filtered_answers = Answers3.objects.filter(Q(user__username__icontains=q), topic=topic).select_related('user', 'question')
    questions = list(Questions.objects.filter(topic=topic))

    def generate_header(questions):
        headers = ['nama siswa']
        for i, question in enumerate(questions, start=1):
            headers.append(f"soal nomer {i}: {question.question}")
            headers.append(f"Umpan Balik soal nomer {i}")
        headers.append('nilai siswa')
        return headers

    headers = generate_header(questions)

    user_answers = defaultdict(lambda: {'answers': {}, 'feedback': {}, 'score': {}})
    for answer in filtered_answers:
        user = answer.user.username
        question_text = answer.question.question
        user_answers[user]['answers'][question_text] = answer.answer
        user_answers[user]['feedback'][question_text] = answer.feedback
        user_answers[user]['score'][question_text] = answer.score or 0

    row_table = []
    for user, data in user_answers.items():
        row_data = [user]
        user_score = 0
        for question in questions:
            row_data.append(data['answers'].get(question.question, ''))
            row_data.append(data['feedback'].get(question.question, ''))
            user_score += data['score'].get(question.question, 0)
        final_score = int(user_score / (3 * len(questions)) * 100)
        row_data.append(final_score)
        row_table.append(row_data)

    context = {'headers': headers, 'row_data': row_table, 'topic': topic}
    return render(request, 'appBelajar/studentsAnswer3.html', context)

@user_passes_test(lambda u: u.is_superuser)
def downloadAnswer1(request, pk):
    topic = get_object_or_404(Topics, id=pk)
    filtered_answers = Answers1.objects.filter(topic=topic).select_related('user', 'question')
    questions = list(Questions.objects.filter(topic=topic))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    def generate_header(questions):
        headers = ['nama siswa']
        for i, question in enumerate(questions, start=1):
            headers.append(f"soal nomer {i}: {question.question}")
            headers.append(f"Umpan Balik soal nomer {i}")
        headers.append('nilai siswa')
        return headers

    headers = generate_header(questions)

    worksheet.append(headers)

    user_answers = {}
    for answer in filtered_answers:
        user = answer.user.username
        question_text = answer.question.question
        answer_text = answer.answer
        feedback_text = answer.feedback
        user_score = answer.score

        if user not in user_answers:
            user_answers[user] = {'answers': {}, 'feedback': {}, 'score': {}}

        user_answers[user]['answers'][question_text] = answer_text
        user_answers[user]['feedback'][question_text] = feedback_text
        user_answers[user]['score'][question_text] = int(user_score)

    for user, data in user_answers.items():
        row_data = [user]
        user_score = 0
        for question in questions:
            answer_text = data['answers'].get(question.question, '')
            feedback_text = data['feedback'].get(question.question, '')
            user_score += data['score'].get(question.question, 0)
            row_data.append(answer_text)
            row_data.append(feedback_text)

        final_score = int(user_score / (3 * len(questions)) * 100)
        row_data.append(final_score)
        worksheet.append(row_data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Jawaban_siswa_{topic.name}.xlsx'
    workbook.save(response)

    return response
    
@user_passes_test(lambda u: u.is_superuser)
def downloadAnswer2(request, pk):
    topic = get_object_or_404(Topics, id=pk)
    filtered_answers = Answers2.objects.filter(topic=topic).select_related('user', 'question')
    questions = list(Questions.objects.filter(topic=topic))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    def generate_header(questions):
        headers = ['nama siswa']
        for i, question in enumerate(questions, start=1):
            headers.append(f"soal nomer {i}: {question.question}")
            headers.append(f"Umpan Balik soal nomer {i}")
        headers.append('nilai siswa')
        return headers

    headers = generate_header(questions)

    worksheet.append(headers)

    user_answers = {}
    for answer in filtered_answers:
        user = answer.user.username
        question_text = answer.question.question
        answer_text = answer.answer
        feedback_text = answer.feedback
        user_score = answer.score

        if user not in user_answers:
            user_answers[user] = {'answers': {}, 'feedback': {}, 'score': {}}

        user_answers[user]['answers'][question_text] = answer_text
        user_answers[user]['feedback'][question_text] = feedback_text
        user_answers[user]['score'][question_text] = int(user_score)

    for user, data in user_answers.items():
        row_data = [user]
        user_score = 0
        for question in questions:
            answer_text = data['answers'].get(question.question, '')
            feedback_text = data['feedback'].get(question.question, '')
            user_score += data['score'].get(question.question, 0)
            row_data.append(answer_text)
            row_data.append(feedback_text)

        final_score = int(user_score / (3 * len(questions)) * 100)
        row_data.append(final_score)
        worksheet.append(row_data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Jawaban_siswa_{topic.name}.xlsx'
    workbook.save(response)

    return response

@user_passes_test(lambda u: u.is_superuser)
def downloadAnswer3(request, pk):
    topic = get_object_or_404(Topics, id=pk)
    filtered_answers = Answers3.objects.filter(topic=topic).select_related('user', 'question')
    questions = list(Questions.objects.filter(topic=topic))

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    def generate_header(questions):
        headers = ['nama siswa']
        for i, question in enumerate(questions, start=1):
            headers.append(f"soal nomer {i}: {question.question}")
            headers.append(f"Umpan Balik soal nomer {i}")
        headers.append('nilai siswa')
        return headers

    headers = generate_header(questions)

    worksheet.append(headers)

    user_answers = {}
    for answer in filtered_answers:
        user = answer.user.username
        question_text = answer.question.question
        answer_text = answer.answer
        feedback_text = answer.feedback
        user_score = answer.score

        if user not in user_answers:
            user_answers[user] = {'answers': {}, 'feedback': {}, 'score': {}}

        user_answers[user]['answers'][question_text] = answer_text
        user_answers[user]['feedback'][question_text] = feedback_text
        user_answers[user]['score'][question_text] = int(user_score)

    for user, data in user_answers.items():
        row_data = [user]
        user_score = 0
        for question in questions:
            answer_text = data['answers'].get(question.question, '')
            feedback_text = data['feedback'].get(question.question, '')
            user_score += data['score'].get(question.question, 0)
            row_data.append(answer_text)
            row_data.append(feedback_text)

        final_score = int(user_score / (3 * len(questions)) * 100)
        row_data.append(final_score)
        worksheet.append(row_data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Jawaban_siswa_{topic.name}.xlsx'
    workbook.save(response)

    return response
